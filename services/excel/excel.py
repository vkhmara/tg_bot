import datetime
from functools import wraps
from typing import Optional, NamedTuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from services.settings import Settings


class ProjectInfo(BaseModel):
    title: str
    ticket: str

    def get_sheetname(self):
        return f"{{{self.ticket}}} {self.title}"


MAX_NOTE_LENGTH = 50


class ProjectNote(NamedTuple):
    note: Optional[str]
    status: Optional[str]
    extra_info: Optional[str]
    created_at: Optional[datetime.datetime | str]
    row: int

    @property
    def cut_note(self) -> str:
        if not self.note:
            return f"{self.row}:"
        if len(self.note) <= MAX_NOTE_LENGTH:
            return f"{self.row}: {self.note}"
        return f"{self.row}: {self.note[:MAX_NOTE_LENGTH]}..."


NEW_NOTE_ROW_NUMBER = 2


class ExcelSheet:
    def __init__(self, sheet: Worksheet):
        self.sheet = sheet

    def get_pending_notes(self) -> list[ProjectNote]:
        notes = []
        for i, sheet_row in enumerate(
            self.sheet.iter_rows(
                max_col=4,
                values_only=True,
            )
        ):
            note = ProjectNote(
                *sheet_row,
                row=i + 1,
            )
            if note.status and note.status.lower() == "pending":
                notes.append(note)
        return notes

    def resolve_note(
        self,
        note: str,
        note_row: int,
    ):
        status_cell = self.sheet.cell(
            row=note_row,
            column=2,
        )
        status_cell.value = "Completed"

    def insert_row(
        self,
        array: list,
        idx: int = NEW_NOTE_ROW_NUMBER,
    ):
        self.sheet.insert_rows(idx=idx)
        for i, value in enumerate(array):
            self.sheet.cell(
                row=idx,
                column=i + 1,
                value=value,
            )


class ExcelTable:
    def __init__(self, document_path: str):
        self._document_path = document_path
        self.wb = load_workbook(document_path)

    def get_sheet(self, sheetname: str) -> ExcelSheet:
        return ExcelSheet(self.wb[sheetname])

    def create_sheet(self, sheetname: str) -> ExcelSheet:
        worksheet = self.wb.create_sheet(title=sheetname)
        return ExcelSheet(sheet=worksheet)

    def get_or_create_sheet(self, sheetname: str) -> ExcelSheet:
        if sheetname not in self.wb.sheetnames:
            return self.create_sheet(sheetname)
        return self.get_sheet(sheetname)

    def copy(self, worksheet: Worksheet) -> Worksheet:
        return self.wb.copy_worksheet(worksheet)

    def save(self):
        self.wb.save(filename=self._document_path)


class Projects:
    PROJECTS_SHEETNAME = "Projects"
    PROJECTS_SHEET_HEADER = ["Project name", "Project ticket(-s)"]
    PROJECT_EXAMPLE_SHEETNAME = "project example"

    def __autosave(func):
        @wraps(func)
        def inner(self, *args, **kwargs):
            return_value = func(self, *args, **kwargs)
            if hasattr(self, "_save"):
                self._save()
            return return_value

        return inner

    def __init__(self):
        self.refresh_excel_table()

    def refresh_excel_table(self):
        settings = Settings()
        self._excel_table = ExcelTable(document_path=settings.settings.document_path)

    def _save(self):
        self._excel_table.save()

    def _get_example_project_sheet(self) -> ExcelSheet:
        return self._excel_table.get_sheet(self.PROJECT_EXAMPLE_SHEETNAME)

    @property
    def all_projects_sheet(self) -> ExcelSheet:
        return self._excel_table.get_sheet(self.PROJECTS_SHEETNAME)

    def get_all_projects_info(self) -> list[ProjectInfo]:
        cell_tuples = list(
            self.all_projects_sheet.sheet.iter_rows(min_row=2, max_col=2)
        )
        return [
            ProjectInfo(
                title=cell_tuple[0].value,
                ticket=cell_tuple[1].value,
            )
            for cell_tuple in cell_tuples
            if cell_tuple[0].value and cell_tuple[1].value
        ]

    def get_all_project_sheetnames(self) -> list[str]:
        return [project.get_sheetname() for project in self.get_all_projects_info()]

    def find_project_sheet_by_name(self, project_name: str) -> Optional[ExcelSheet]:
        project_sheetname = next(
            (
                sheetname
                for sheetname in self.get_all_project_sheetnames()
                if project_name.lower() in sheetname.lower()
            ),
            None,
        )
        if not project_sheetname:
            print(f"No Ticket In Projects Sheet: {project_name}")
            return
        return self._excel_table.get_or_create_sheet(project_name)

    def has_project(self, project_name: str) -> bool:
        return any(
            project_name.lower() in sheetname.lower()
            for sheetname in self.get_all_project_sheetnames()
        )

    def create_project_sheet(
        self,
        sheetname: str,
        commit: bool = False,
    ) -> Worksheet:
        example_project_sheet = self._get_example_project_sheet()
        worksheet = self._excel_table.copy(example_project_sheet.sheet)
        worksheet.title = sheetname
        if commit:
            self._save()
        return worksheet

    @__autosave
    def validate_sheets(self):
        sheetnames = self._excel_table.wb.sheetnames
        for project_info in self.get_all_projects_info():
            project_sheetname = project_info.get_sheetname()
            if project_sheetname not in sheetnames:
                self.create_project_sheet(project_sheetname)

    @__autosave
    def create_note_sheet(self, ticket: str) -> Optional[ExcelSheet]:
        sheets = [
            x
            for x in self.get_all_projects_info()
            if ticket.lower() in x.get_sheetname().lower()
        ]
        if not sheets:
            print(f"No Ticket In Projects Sheet: {ticket}")
            return
        if len(sheets) > 1:
            print(f"Multiple Tickets In Projects Sheet: {ticket}")
        sheetname = sheets[0].get_sheetname()
        excel_sheet = self.create_project_sheet(
            sheetname=sheetname,
        )
        return ExcelSheet(excel_sheet)

    @__autosave
    def add_note(
        self,
        ticket: str,
        note: str,
        extra_info: str = None,
    ):
        if not ticket:
            raise ValueError(f"`ticket` Cannot Be Falsish")
        sheetnames = [
            x for x in self._excel_table.wb.sheetnames if ticket.lower() in x.lower()
        ]
        if sheetnames:
            sheetname = sheetnames[0]
            excel_sheet = self._excel_table.get_sheet(sheetname)
        else:
            excel_sheet = self.create_note_sheet(ticket)

        if not excel_sheet:
            raise Exception(f"Excel Sheet Wasn't Created: {ticket=}")

        excel_sheet.insert_row(
            array=[
                note,
                "Pending",
                extra_info or "",
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        )

    @__autosave
    def add_project(
        self,
        ticket: str,
        ticket_name: str,
    ):
        self.all_projects_sheet.insert_row(array=[ticket_name, ticket])

    def get_pending_notes(self, project_name: str) -> list[ProjectNote]:
        project_excel_sheet = self.find_project_sheet_by_name(project_name=project_name)
        if not project_excel_sheet:
            raise Exception(f"No such project {project_name}")
        return project_excel_sheet.get_pending_notes()

    @__autosave
    def resolve_note(
        self,
        project_name: str,
        note: str,
        note_row: int,
    ):
        project_excel_sheet = self.find_project_sheet_by_name(project_name=project_name)
        if not project_excel_sheet:
            raise Exception(f"No such project {project_name}")
        project_excel_sheet.resolve_note(
            note=note,
            note_row=note_row,
        )
